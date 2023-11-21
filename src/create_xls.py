import json
import os

import openpyxl


# Загружаем данные из JSON файла
filepath = os.path.join("data", "file_finish.json")
with open(filepath, "r", encoding="UTF-8") as file:
    users_data = json.load(file)

# Создаем новый файл Excel
workbook = openpyxl.Workbook()

# Итерируемся по данным и создаем листы для каждой даты
for date, data in users_data.items():
    if date != 'total_tasks':
        # Создаем новый лист для текущей даты
        sheet = workbook.create_sheet(title=date)

        # Заполняем заголовки столбцов
        sheet['A1'] = 'ФИО'
        sheet['B1'] = 'Количество'

        # Заполняем лист данными
        row_index = 2
        for employee, tasks_count in data['employees'].items():
            sheet[f'A{row_index}'] = employee
            sheet[f'B{row_index}'] = tasks_count
            row_index += 1

# Создаем лист "ИТОГ"
total_sheet = workbook.create_sheet(title="ИТОГ")

# Заполняем заголовки столбцов в листе "ИТОГ"
total_sheet['A1'] = 'ФИО'
total_sheet['B1'] = 'Общее количество'

# Инициализируем словарь для хранения общего количества задач по пользователям
total_tasks_by_employee = {}

# Итерируемся по данным и суммируем общее количество задач по пользователям
for date, data in users_data.items():
    if date != 'total_tasks':
        employees_data = data['employees']
        for employee, tasks_count in employees_data.items():
            if employee not in total_tasks_by_employee:
                total_tasks_by_employee[employee] = 0
            total_tasks_by_employee[employee] += tasks_count

# Заполняем лист "ИТОГ" данными
row_index = 2
for employee, total_tasks in total_tasks_by_employee.items():
    total_sheet[f'A{row_index}'] = employee
    total_sheet[f'B{row_index}'] = total_tasks
    row_index += 1

# Удаляем лист по умолчанию
workbook.remove(workbook['Sheet'])

# Сохраняем файл Excel
workbook.save("output_excel_file.xlsx")
