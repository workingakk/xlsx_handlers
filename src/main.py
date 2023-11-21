import json
import os

import openpyxl


def process_excel_file(file_path):
    # Загрузка файла Excel
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active

    # Создание словаря для хранения данных о сотрудниках
    users_data = {}

    # Сортировка данных по дате закрытия задач
    rows = list(sheet.iter_rows(min_row=2, max_col=sheet.max_column, max_row=sheet.max_row, values_only=True))
    sorted_rows = sorted(rows, key=lambda x: x[7])

    # Обработка отсортированных данных
    for row in sorted_rows:
        if len(row) == 8:
            _, _, _, executor, _, _, _, close_date = row
            format_close_date = close_date.strftime("%m.%Y")
            if users_data.get(format_close_date) is None:
                users_data[format_close_date] = {'total_tasks': 0}

            # Подсчет общего количества задач
            users_data[format_close_date]['total_tasks'] += 1

            # Подсчет общего количества задач для каждого сотрудника
            if executor not in users_data[format_close_date]:
                users_data[format_close_date][executor] = 0
            users_data[format_close_date][executor] += 1


    # Добавление блока с общим количеством задач и подсчетом общей суммы закрытых задач каждым сотрудником
    for date, data in users_data.items():
        if date != 'total_tasks':
            total_tasks = data['total_tasks']
            employee_data = {employee: tasks_count for employee, tasks_count in data.items() if
                             employee != 'total_tasks'}
            sorted_employee_data = dict(sorted(employee_data.items(), key=lambda item: item[1], reverse=True))
            users_data[date] = {'total_tasks': total_tasks, 'employees': sorted_employee_data}

    with open("file.json", "w", encoding="UTF-8") as file:
        json.dump(users_data, file, ensure_ascii=False, indent=4)


    # Вывод результатов
    print("Общее количество задач:", users_data[format_close_date].get('total_tasks', 0))
    print("Общее количество задач по сотрудникам:")
    for date, data in users_data.items():
        if date != 'total_tasks':
            print(f"{date}: {data['total_tasks']}")
            print("Закрытые задачи по сотрудникам:")
            for employee, tasks_count in data['employees'].items():
                print(f"  {employee}: {tasks_count}")

    # Закрытие файла Excel
    workbook.close()


filepath = os.path.join("data", "all_tasks.xlsx")
print(os.path.isfile(filepath))
process_excel_file(filepath)
