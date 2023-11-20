import openpyxl
from datetime import datetime

import os

filepath = os.path.join("data", "test.xlsx")


def process_excel_file(file_path):
    # Загрузка файла Excel
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active

    # Создание словаря для хранения данных о сотрудниках
    data = {}

    # Сортировка данных по дате закрытия задач
    rows = list(sheet.iter_rows(min_row=2, max_col=sheet.max_column, max_row=sheet.max_row, values_only=True))
    #sorted_rows = sorted(rows, key=lambda x: x[7])# if isinstance(x[7], str) else x[7].strftime('%d.%m.%Y %H:%M'))

    #  Обработка отсортированных данных
    for row in rows:
        if len(row) == 8:
            _, _, _, executor, _, _, _, close_date = row
            #close_date = close_date if isinstance(close_date, str) else close_date.strftime('%d.%m.%Y %H:%M')
            format_close_date = close_date.strftime("%m.%Y")
            print(format_close_date)
            print(data.get(format_close_date, False))
            if not data.get(format_close_date, False):
                print(1)
            quit()
            if data.get(format_close_date, False):
                data[format_close_date] = {}
            # Подсчет общего количества задач
            data[format_close_date]['total_tasks'] = data.get('total_tasks', 0) + 1

            # Подсчет общего количества задач для каждого сотрудника
            data[format_close_date][executor] = data.get(executor, 0) + 1
            #break
        print(data)
    # Вывод результатов
    # print("Общее количество задач:", data.get('total_tasks', 0))
    # print("Общее количество задач по сотрудникам:")
    # for employee, tasks_count in data.items():
    #     if employee != 'total_tasks':
    #         print(f"{employee}: {tasks_count}")

    # Закрытие файла Excel
    workbook.close()

# Пример использования
#excel_file_path = 'your_excel_file.xlsx'
process_excel_file(filepath)
